import os
import sys
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import openpyxl
import scipy.io.wavfile as wav
from python_speech_features import *
from util.sound_util import *
from pylab import *
import pandas as pd
from openpyxl.styles import PatternFill


def extract_mfcc_feat(input_dir):
    (rate, sig) = wav.read(input_dir)
    signal = end_point_detection(input_dir)
    for i in range(len(signal)):
        signal[i] = int(round(signal[i]))
    signal = np.array(signal)
    frame_len = 0.02
    frame_step = 0.01
    frame_size = int(round(frame_len * rate))
    mfcc_feat = mfcc(signal, samplerate=rate, winlen=frame_len, winstep=frame_step, nfft=frame_size)
    return mfcc_feat


def get_all_mfcc_feat():
    training_data = []
    test_data = []
    training_path = "../data/training data"
    test_path = "../data/test data"
    for file in os.listdir(training_path):
        if file.split('.')[1] == "wav":
            training_data.append(extract_mfcc_feat(os.path.join(training_path, file)))
    for file in os.listdir(test_path):
        if file.split('.')[1] == "wav":
            test_data.append(extract_mfcc_feat(os.path.join(test_path, file)))

    mfcc_dynamic_programming(training_data[5], test_data[5])
    # con_table = zeros((6, 6))
    # for i in range(6):
    #     for j in range(6):
    #         con_table[i][j] = mfcc_dynamic_programming(training_data[i], test_data[j])
    # print(con_table)


def mfcc_dynamic_programming(mfcc1, mfcc2):
    col = mfcc1.shape[0]
    row = mfcc2.shape[0]
    dp_matrix = zeros((col, row))
    restricted_col = int(round(col * 0.8))
    restricted_row = int(round(row * 0.8))

    restricted_table = zeros((col, row))

    for i in range(col):
        for j in range(row):
            for k in range(1, 13):
                dp_matrix[i][j] += pow(mfcc1[i, k] - mfcc2[j, k], 2)
            dp_matrix[i][j] = sqrt(dp_matrix[i][j])

    accumulate_matrix = zeros((col, row))
    accumulate_matrix[0, 0] = dp_matrix[0, 0]
    for i in range(col):
        for j in range(row):
            if i == 0:
                accumulate_matrix[i, j] = dp_matrix[i, j] + accumulate_matrix[i, j - 1]
            elif j == 0:
                accumulate_matrix[i, j] = dp_matrix[i, j] + accumulate_matrix[i - 1, j]
            else:
                accumulate_matrix[i, j] = dp_matrix[i, j] + min(accumulate_matrix[i, j - 1],
                                                                accumulate_matrix[i - 1, j - 1],
                                                                accumulate_matrix[i - 1, j])
    min_value = +inf
    min_index = [0, 0]
    for i in range(restricted_row, row):
        if accumulate_matrix[col - 1][i] < min_value:
            min_value = accumulate_matrix[col - 1][i]
            min_index = [col - 1, i]
    for i in range(restricted_col, col):
        if accumulate_matrix[i][row - 1] < min_value:
            min_value = accumulate_matrix[i][row - 1]
            min_index = [i, row - 1]
    # print(min_index)
    # print(min_value)
    optimal_path = [min_index]
    current_index = min_index
    optimal_list = [min_value]
    i = min_index[0]
    j = min_index[1]
    while current_index[0] > 0 and current_index[1] > 0:
        next_choices = [accumulate_matrix[current_index[0] - 1][current_index[1] - 1],
                        accumulate_matrix[current_index[0]][current_index[1] - 1],
                        accumulate_matrix[current_index[0] - 1][current_index[1]]]
        smallest_index = next_choices.index(min(next_choices))
        if smallest_index == 0:
            current_index = [current_index[0] - 1, current_index[1] - 1]
        elif smallest_index == 1:
            current_index = [current_index[0], current_index[1] - 1]
        else:
            current_index = [current_index[0] - 1, current_index[1]]
        optimal_path.append(current_index)
        optimal_list.append(next_choices[smallest_index])
    tmp = []
    for i in range(len(optimal_path)):
        tmp.append(optimal_path[len(optimal_path) - i - 1])
    optimal_path = tmp
    print(optimal_path)

    paint_dp_matrix(accumulate_matrix, optimal_path)
    return min_value


# Write the dp matrix into excel file and paint the bg color.
def paint_dp_matrix(dp_matrix, optimal_path):
    data = []
    for i in range(dp_matrix.shape[0]):
        line = []
        for j in range(dp_matrix.shape[1]):
            line.append(dp_matrix[dp_matrix.shape[0] - i - 1][j])
        data.append(line)
    data = np.array(data)
    data_df = pd.DataFrame(data)
    columns = []
    rows = []
    for i in range(data.shape[0]):
        columns.append(data.shape[0] - i)
    for i in range(data.shape[1]):
        rows.append(i + 1)
    for i in range(len(optimal_path)):
        optimal_path[i][0] = data.shape[0] - optimal_path[i][0]
        optimal_path[i][1] += 1
    data_df.columns = rows
    data_df.index = columns
    writer = pd.ExcelWriter('accumulate_dp_matrix.xlsx')
    data_df.to_excel(writer, float_format='%.5f')
    writer.save()
    # paint the bg color
    excel_file = "accumulate_dp_matrix.xlsx"
    ff = openpyxl.load_workbook(excel_file)
    fill = PatternFill("solid", start_color="FFFF00")
    for idx, col in enumerate(ff.worksheets[0]):
        for idx2, row in enumerate(col):
            if [idx, idx2] in optimal_path:
                row.fill = fill
    ff.save('accumulate_dp_matrix.xlsx')

if __name__ == '__main__':
    # extract_mfcc_feat("C:\\Users\\24111\\PycharmProjects\\voice-recognition\\data\\training data\\s1a.wav")
    get_all_mfcc_feat()
